import openpyxl
import os

# os.chdir('C:\\Users\\praju\\')

workbook = openpyxl.load_workbook('readExcel.xlsx')
# print(type(workbook))
sheet = workbook.get_sheet_by_name('Sheet1')
# print(type(sheet))
# print(workbook.get_sheet_names())
# print(sheet['A1'])
cell1 = sheet['A1']
cell2 = sheet['B4']
print(cell1.value)
print(cell2.value)
print(sheet.cell(row=1, column=2).value)


wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'].value = 42
os.chdir('E:\\DATA\\Projects\\Python Projects\\Automate The Boring Stuff with Python\\Programs')
wb.save('writeExcel.xlsx')
newSheet = wb.create_sheet()
newSheet.title = 'Fresh Sheet'
wb.create_sheet(index=0, title="Other new sheet")
